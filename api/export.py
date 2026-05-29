"""
api/export.py
──────────────
New endpoints — register these in your api/routes.py or api/main.py:

    from api.export import router as export_router
    app.include_router(export_router)

Endpoints:
    GET  /documents/{id}/export/csv           → download CSV file
    GET  /documents/{id}/export/excel         → download Excel (.xlsx) file
    GET  /documents/{id}/export/email?to=x@y  → send Excel by email
    GET  /stats                               → live counter (total docs + avg confidence)
"""

from __future__ import annotations

import contextlib
import csv
import io
import json
import os
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.orm import Session

from auth.core import get_current_user
from database.connection import get_db_for_fastapi
from database.models import Document, ExtractedField, User

router = APIRouter(tags=["export"])


# ─────────────────────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────────────────────


def _get_doc_and_fields(doc_id: int, db: Session, user: User):
    """Fetch doc + fields, enforcing ownership. Returns 404 if not found or not owned."""
    doc = (
        db.query(Document)
        .filter(Document.id == doc_id, Document.user_id == user.id)
        .first()
    )
    if not doc:
        # Deliberately 404 not 403: don't reveal that the document exists
        raise HTTPException(status_code=404, detail="Document not found")
    fields = db.query(ExtractedField).filter(ExtractedField.document_id == doc_id).all()
    return doc, fields


def _scalar_rows(fields) -> dict:
    """Return all non-line-item fields as a flat dict."""
    return {
        f.field_name: f.field_value for f in fields if not f.field_name.startswith("line_items_")
    }


def _line_item_rows(fields) -> list[dict]:
    """Parse all line_items_* fields into a list of dicts."""
    items = []
    for f in sorted(fields, key=lambda x: x.field_name):
        if f.field_name.startswith("line_items_"):
            with contextlib.suppress(json.JSONDecodeError, TypeError):
                items.append(json.loads(f.field_value))
    return items


# ─────────────────────────────────────────────────────────────
#  CSV export
# ─────────────────────────────────────────────────────────────


@router.get("/documents/{doc_id}/export/csv")
def export_csv(
    doc_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db_for_fastapi),
):
    """Download extracted invoice data as a CSV file (only for documents you own)."""
    doc, fields = _get_doc_and_fields(doc_id, db, current_user)
    scalars = _scalar_rows(fields)
    line_items = _line_item_rows(fields)

    output = io.StringIO()
    writer = csv.writer(output)

    # ── Invoice summary section ──
    writer.writerow(["INVOICE SUMMARY"])
    writer.writerow(["Field", "Value"])
    summary_keys = [
        "vendor_name",
        "vendor_gstin",
        "buyer_name",
        "buyer_gstin",
        "invoice_number",
        "invoice_date",
        "due_date",
        "currency",
        "subtotal",
        "tax_amount",
        "total_amount",
        "bank_name",
        "bank_ifsc",
        "bank_account_number",
    ]
    for key in summary_keys:
        val = scalars.get(key, "")
        if val:
            writer.writerow([key.replace("_", " ").title(), val])

    writer.writerow([])
    writer.writerow(["AI Confidence", f"{doc.ai_confidence:.0%}" if doc.ai_confidence else ""])

    # ── Line items section ──
    if line_items:
        writer.writerow([])
        writer.writerow(["LINE ITEMS"])
        writer.writerow(["#", "Description", "HSN", "Qty", "Unit", "Unit Price", "Amount"])
        for i, item in enumerate(line_items, start=1):
            writer.writerow(
                [
                    i,
                    item.get("description", ""),
                    item.get("hsn", ""),
                    item.get("quantity", ""),
                    item.get("unit", ""),
                    item.get("unit_price", ""),
                    item.get("amount", ""),
                ]
            )

    csv_bytes = output.getvalue().encode("utf-8-sig")  # utf-8-sig adds BOM for Excel compatibility
    filename = f"invoice_{doc.invoice_number or doc_id}.csv".replace("/", "-")

    return Response(
        content=csv_bytes,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────
#  Excel export
# ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT = Font(bold=True, size=10)
ZEBRA_FILL = PatternFill("solid", fgColor="EBF3FB")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _build_excel(doc, scalars: dict, line_items: list[dict]) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Invoice Details ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Invoice Details"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 42

    # Title row
    ws1.merge_cells("A1:B1")
    title_cell = ws1["A1"]
    title_cell.value = "INVOICE EXTRACTION REPORT"
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = CENTER
    ws1.row_dimensions[1].height = 28

    rows = [
        ("Vendor Name", scalars.get("vendor_name")),
        ("Vendor GSTIN", scalars.get("vendor_gstin")),
        ("Buyer Name", scalars.get("buyer_name")),
        ("Buyer GSTIN", scalars.get("buyer_gstin")),
        ("Invoice Number", scalars.get("invoice_number")),
        ("Invoice Date", scalars.get("invoice_date")),
        ("Due Date", scalars.get("due_date")),
        ("Currency", scalars.get("currency")),
        ("Subtotal", scalars.get("subtotal")),
        ("Tax Amount", scalars.get("tax_amount")),
        ("Total Amount", scalars.get("total_amount")),
        ("Bank Name", scalars.get("bank_name")),
        ("Bank IFSC", scalars.get("bank_ifsc")),
        ("Bank Account Number", scalars.get("bank_account_number")),
        ("AI Confidence", f"{doc.ai_confidence:.0%}" if doc.ai_confidence else ""),
    ]

    for r, (label, value) in enumerate(rows, start=2):
        if not value:
            continue
        label_cell = ws1.cell(row=r, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.alignment = LEFT
        value_cell = ws1.cell(row=r, column=2, value=value)
        value_cell.alignment = LEFT
        if r % 2 == 0:
            value_cell.fill = ZEBRA_FILL

    # ── Sheet 2: Line Items ───────────────────────────────────
    ws2 = wb.create_sheet("Line Items")
    headers = ["#", "Description", "HSN Code", "Quantity", "Unit", "Unit Price", "Amount"]
    col_widths = [5, 50, 12, 10, 10, 14, 16]

    for col, (h, w) in enumerate(zip(headers, col_widths, strict=True), start=1):
        ws2.column_dimensions[get_column_letter(col)].width = w
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = SUBHEAD_FONT
        cell.fill = SUBHEAD_FILL
        cell.alignment = CENTER

    for r, item in enumerate(line_items, start=2):
        row_data = [
            r - 1,
            item.get("description", ""),
            item.get("hsn", ""),
            item.get("quantity", ""),
            item.get("unit", ""),
            item.get("unit_price", ""),
            item.get("amount", ""),
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r, column=col, value=val)
            cell.alignment = LEFT
            if r % 2 == 0:
                cell.fill = ZEBRA_FILL

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/documents/{doc_id}/export/excel")
def export_excel(
    doc_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db_for_fastapi),
):
    """Download extracted invoice data as a formatted Excel (.xlsx) file (only for documents you own)."""
    doc, fields = _get_doc_and_fields(doc_id, db, current_user)
    scalars = _scalar_rows(fields)
    line_items = _line_item_rows(fields)

    xlsx_bytes = _build_excel(doc, scalars, line_items)
    filename = f"invoice_{doc.invoice_number or doc_id}.xlsx".replace("/", "-")

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────
#  Email export
# ─────────────────────────────────────────────────────────────


@router.get("/documents/{doc_id}/export/email")
def export_email(
    doc_id: int,
    to: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db_for_fastapi),
):
    """
    Send the Excel file to the given email address (only for documents you own).
    Requires these env vars:
        SMTP_HOST     e.g. smtp.gmail.com
        SMTP_PORT     e.g. 587
        SMTP_USER     your sending email address
        SMTP_PASSWORD your app password (Gmail: create at myaccount.google.com/apppasswords)
    """
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER")
    smtp_pass = os.getenv("SMTP_PASSWORD")

    if not all([smtp_host, smtp_user, smtp_pass]):
        raise HTTPException(
            status_code=503,
            detail=(
                "Email not configured. Set SMTP_HOST, SMTP_USER, SMTP_PASSWORD "
                "in Render environment variables."
            ),
        )

    doc, fields = _get_doc_and_fields(doc_id, db, current_user)
    scalars = _scalar_rows(fields)
    line_items = _line_item_rows(fields)
    xlsx_bytes = _build_excel(doc, scalars, line_items)
    filename = f"invoice_{doc.invoice_number or doc_id}.xlsx".replace("/", "-")

    # Build email
    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = to
    msg["Subject"] = f"Invoice Data — {doc.invoice_number or f'Document #{doc_id}'}"
    msg["X-Priority"] = "3"

    body = MIMEText(
        f"Please find the extracted invoice data attached.\n\n"
        f"Vendor:  {scalars.get('vendor_name', 'N/A')}\n"
        f"Invoice: {scalars.get('invoice_number', 'N/A')}\n"
        f"Date:    {scalars.get('invoice_date', 'N/A')}\n"
        f"Total:   {scalars.get('currency', 'INR')} {scalars.get('total_amount', 'N/A')}\n\n"
        f"AI Confidence: {doc.ai_confidence:.0%}\n\n"
        f"Extracted by AI Document Intelligence.",
        "plain",
    )
    msg.attach(body)

    attachment = MIMEBase("application", "octet-stream")
    attachment.set_payload(xlsx_bytes)
    encoders.encode_base64(attachment)
    attachment.add_header("Content-Disposition", f'attachment; filename="{filename}"')
    msg.attach(attachment)

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, to, msg.as_string())
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email send failed: {e}") from e

    return {"success": True, "sent_to": to, "filename": filename}


# ─────────────────────────────────────────────────────────────
#  Live stats — global counter + average confidence
# ─────────────────────────────────────────────────────────────


@router.get("/stats")
def get_platform_stats(db: Session = Depends(get_db_for_fastapi)):
    """
    Returns live platform-wide stats for the trust counter widget.
    Response:
        {
          "total_documents": 142,
          "avg_confidence": 0.91,
          "avg_confidence_pct": "91%"
        }
    """
    try:
        row = db.execute(
            text("SELECT total_documents, confidence_sum FROM platform_stats WHERE id = 1")
        ).fetchone()

        if not row or row[0] == 0:
            return {"total_documents": 0, "avg_confidence": 0.0, "avg_confidence_pct": "0%"}

        total = row[0]
        avg = round(row[1] / total, 4)
        return {
            "total_documents": total,
            "avg_confidence": avg,
            "avg_confidence_pct": f"{avg:.0%}",
        }
    except Exception:
        # Table might not exist yet — return zeros gracefully
        return {"total_documents": 0, "avg_confidence": 0.0, "avg_confidence_pct": "0%"}
